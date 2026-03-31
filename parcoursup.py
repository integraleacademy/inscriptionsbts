# parcoursup.py
import os, uuid, sqlite3, json, re, time
import requests
from flask import current_app  # <— nouveau
import sys                     # <— nouveau
sys.stdout.reconfigure(line_buffering=True)
from datetime import datetime
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from utils import send_mail, send_sms_brevo
from mail_templates import mail_html
from sms_templates import sms_text
from openpyxl import load_workbook
import unicodedata
from markupsafe import Markup

def normalize(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


# Blueprint Parcoursup
bp_parcoursup = Blueprint("parcoursup", __name__, template_folder="templates")

DATA_DIR = os.getenv("DATA_DIR", "/data")
DB_PATH = os.path.join(DATA_DIR, "app.db")
CLEANED_DIR = os.path.join(DATA_DIR, "parcoursup_cleaned")
os.makedirs(CLEANED_DIR, exist_ok=True)

def db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    return conn

def _clean_phone(raw_tel):
    tel = str(raw_tel or "").strip().replace(" ", "")
    tel = tel.replace(".", "").replace("-", "")
    if tel.startswith("0033"):
        tel = "+33" + tel[4:]
    elif re.match(r"^33[1-9]\d{8}$", tel):
        tel = "+33" + tel[2:]
    elif re.match(r"^0[1-9]\d{8}$", tel):
        tel = "+33" + tel[1:]
    return tel

def _clean_mode(raw_mode):
    mode = (raw_mode or "").strip().lower()
    if mode in ("presentiel", "présentiel"):
        return "Présentiel"
    if mode == "distanciel":
        return "Distanciel"
    return raw_mode

def get_stats_parcoursup():
    conn = db()
    cur = conn.cursor()

    stats = {
        "total": cur.execute("SELECT COUNT(*) FROM parcoursup_candidats").fetchone()[0],
        "mail_sent": 0,
        "mail_delivered": 0,
        "mail_opened": 0,
        "mail_clicked": 0,
        "sms_sent": 0,
        "sms_delivered": 0,
    }

    # On lit tous les logs JSON
    rows = cur.execute("SELECT logs FROM parcoursup_candidats").fetchall()
    for (raw_logs,) in rows:
        try:
            logs = json.loads(raw_logs or "[]")
            seen = set()  # ⚡ compteur unique par candidat

            for log in logs:
                if log.get("type") == "mail":
                    stats["mail_sent"] += 1
                elif log.get("type") == "mail_status":
                    evt = log.get("event")
                    if evt == "delivered" and "delivered" not in seen:
                        stats["mail_delivered"] += 1
                        seen.add("delivered")
                    elif evt in ["opened", "unique_opened", "click"]:
                        # 🧩 Gestion propre des statuts mail
                        if "opened" not in seen:
                            stats["mail_opened"] += 1
                            seen.add("opened")
                        if evt == "click" and "click" not in seen:
                            stats["mail_clicked"] += 1
                            seen.add("click")

                elif log.get("type") == "sms":
                    stats["sms_sent"] += 1
                elif log.get("type") == "sms_status" and log.get("event") == "delivered":
                    stats["sms_delivered"] += 1

        except Exception:
            continue

    conn.close()
    return stats




STATUTS_STYLE = {
    "preinscription": {"label": "Pré-inscription à traiter", "color": "#808080"},
    "validee": {"label": "Candidature validée", "color": "#3498db"},
    "confirmee": {"label": "Inscription confirmée", "color": "#f4c45a"},
    "reconf_en_cours": {"label": "Reconfirmation en cours", "color": "#ff9800"},
    "reconfirmee": {"label": "Inscription re-confirmée", "color": "#2ecc71"},
    "annulee": {"label": "Inscription annulée", "color": "#e74c3c"},
    "docs_non_conformes": {"label": "Documents non conformes", "color": "#000000"},
    "En attente de candidature": {"label": "En attente de candidature", "color": "#e74c3c"},
}

# =====================================================
# 🧱 Initialisation table Parcoursup
# =====================================================
def init_parcoursup_table():
    conn = db()
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS parcoursup_candidats (
        id TEXT PRIMARY KEY,
        nom TEXT,
        prenom TEXT,
        telephone TEXT,
        email TEXT,
        formation TEXT,
        mode TEXT,
        mail_ok INTEGER DEFAULT 0,
        sms_ok INTEGER DEFAULT 0,
        statut TEXT DEFAULT 'En attente de candidature',
        logs TEXT DEFAULT '[]',
        created_at TEXT
    );
    """)
    conn.commit()
    conn.close()

init_parcoursup_table()

# =====================================================
# 🏠 PAGE PRINCIPALE AVEC SYNCHRONISATION AUTO
# =====================================================
@bp_parcoursup.route("/parcoursup")
def dashboard():
    conn = db()
    cur = conn.cursor()

    try:
        cur.execute("SELECT id, email, telephone FROM parcoursup_candidats")
        parcoursup_rows = cur.fetchall()

        for r in parcoursup_rows:
            email = (r["email"] or "").strip().lower()
            tel = (r["telephone"] or "").replace(" ", "").replace("+33", "0").strip()

            cur2 = conn.execute("""
                SELECT statut FROM candidats
                WHERE LOWER(TRIM(email)) = ?
                   OR REPLACE(REPLACE(tel, ' ', ''), '+33', '0') = ?
            """, (email, tel))
            existing = cur2.fetchone()

            if existing and existing["statut"]:
                cur.execute("UPDATE parcoursup_candidats SET statut=? WHERE id=?", (existing["statut"], r["id"]))

        conn.commit()
        print("✅ Synchronisation Parcoursup ↔ Admin réussie")

    except Exception as e:
        print("⚠️ Erreur de synchronisation Parcoursup ↔ Admin :", e)

    # ==========================================================
    # 📦 Chargement avec filtres (statut / formation / mode)
    # ==========================================================
    statut_q = request.args.get("statut", "").strip()
    formation_q = request.args.get("formation", "").strip()
    mode_q = request.args.get("mode", "").strip()
    search_q = request.args.get("search", "").strip()

    where = []
    params = []

    if statut_q:
        where.append("statut = ?")
        params.append(statut_q)

    if formation_q:
        where.append("formation = ?")
        params.append(formation_q)

    if mode_q:
        where.append("LOWER(mode) = LOWER(?)")
        params.append(mode_q)

    if search_q:
        where.append("(LOWER(nom) LIKE ? OR LOWER(prenom) LIKE ? OR LOWER(email) LIKE ?)")
        pattern = f"%{normalize(search_q.lower())}%"
        params.extend([pattern, pattern, pattern])




    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    cur.execute(f"SELECT * FROM parcoursup_candidats {where_sql} ORDER BY created_at DESC", params)
    rows = [dict(r) for r in cur.fetchall()]

    # Pour alimenter les listes déroulantes dans les filtres
    formations = [r[0] for r in cur.execute("SELECT DISTINCT formation FROM parcoursup_candidats ORDER BY formation").fetchall() if r[0]]
    modes = [r[0] for r in cur.execute("SELECT DISTINCT mode FROM parcoursup_candidats ORDER BY mode").fetchall() if r[0]]



    # 🧩 On décode les logs JSON ici (au lieu du |fromjson dans Jinja)
    for r in rows:
        try:
            r["logs"] = json.loads(r.get("logs", "[]"))
        except Exception:
            r["logs"] = []

    # 🔍 Extraction des dates d’envoi Mail/SMS et statuts détaillés
    for r in rows:
        try:
            logs = r.get("logs", [])
            mail_log = next((log for log in logs if log.get("type") == "mail"), None)
            sms_log = next((log for log in logs if log.get("type") == "sms"), None)
            r["mail_date"] = mail_log.get("date") if mail_log else None
            r["sms_date"] = sms_log.get("date") if sms_log else None

            # Détermination du dernier statut mail
            mail_status = next((l.get("event") for l in reversed(logs) if l.get("type") == "mail_status"), None)
            if mail_status in ["click"]:
                r["mail_status_label"] = "Cliqué"
                r["mail_status_color"] = "#8e44ad"  # violet
            elif mail_status in ["opened", "unique_opened"]:
                r["mail_status_label"] = "Ouvert"
                r["mail_status_color"] = "#3498db"  # bleu
            elif mail_status == "delivered":
                r["mail_status_label"] = "Délivré"
                r["mail_status_color"] = "#2ecc71"  # vert
            elif mail_log:
                r["mail_status_label"] = "Envoyé"
                r["mail_status_color"] = "#f1c40f"  # jaune
            else:
                r["mail_status_label"] = "Non envoyé"
                r["mail_status_color"] = "#7f8c8d"

        except Exception:
            r["mail_status_label"] = "Inconnu"
            r["mail_status_color"] = "#7f8c8d"

    # 🔍 Ajout du dernier statut SMS pour affichage direct (priorité au "delivered")
    for r in rows:
        try:
            logs = r.get("logs", [])
            # on cherche d’abord un statut livré
            sms_delivered = next(
                (l for l in logs if l.get("type") == "sms_status" and l.get("event") == "delivered"),
                None
            )
            if sms_delivered:
                r["sms_status"] = "delivered"
            else:
                # sinon on prend le dernier évènement connu
                sms_status = next(
                    (l.get("event") for l in reversed(logs) if l.get("type") == "sms_status"),
                    None
                )
                r["sms_status"] = sms_status or "unknown"
        except Exception:
            r["sms_status"] = "unknown"

        # 🔎 Marquage des candidats sans ouverture/clic après 48h
    now = datetime.now()
    for r in rows:
        try:
            logs = r.get("logs", [])
            mail_log = next((l for l in logs if l.get("type") == "mail"), None)
            opened_or_clicked = any(
                l.get("event") in ["opened", "click", "unique_opened"]
                for l in logs if l.get("type") == "mail_status"
            )
            if mail_log and not opened_or_clicked:
                mail_date = datetime.fromisoformat(mail_log["date"])
                diff_hours = (now - mail_date).total_seconds() / 3600
                if diff_hours > 48:
                    r["alerte"] = True
                else:
                    r["alerte"] = False
            else:
                r["alerte"] = False
        except Exception:
            r["alerte"] = False

    # 🟡 Marquage visuel des candidats relancés
    for r in rows:
        try:
            logs = r.get("logs", [])
            relance = next(
                (l for l in reversed(logs) if l.get("type") in ["relance_manuelle", "relance_auto"]),
                None
            )
            r["relancee"] = bool(relance)
        except Exception:
            r["relancee"] = False
  

    conn.close()

    stats = get_stats_parcoursup()
    return render_template(
        "parcoursup.html",
        title="Gestion Parcoursup",
        rows=rows,
        STATUTS_STYLE=STATUTS_STYLE,
        stats=stats,
        formations=formations,
        modes=modes

    )







# =====================================================
# 📤 IMPORTER UN FICHIER EXCEL (.xlsx) — AVEC LOGS + PAUSES
# =====================================================
@bp_parcoursup.route("/parcoursup/import", methods=["POST"])
def import_file():
    if "file" not in request.files:
        flash("Aucun fichier sélectionné", "error")
        return redirect(url_for("parcoursup.dashboard"))

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        flash("Format non supporté. Utilisez un fichier .xlsx", "error")
        return redirect(url_for("parcoursup.dashboard"))

    temp_path = os.path.join(DATA_DIR, secure_filename(file.filename))
    file.save(temp_path)
    wb = load_workbook(temp_path)
    ws = wb.active

    LOT_SIZE = 100
    total_rows = ws.max_row - 1
    total_lots = (total_rows // LOT_SIZE) + (1 if total_rows % LOT_SIZE else 0)

    print(f"🚀 Import total : {total_rows} lignes ({total_lots} lots de {LOT_SIZE})")

    conn = db()
    cur = conn.cursor()
    imported = mails_sent = sms_sent = duplicates = errors = 0
    row_index = 0

    try:
        for lot_index in range(total_lots):
            start_row = 2 + lot_index * LOT_SIZE
            end_row = min(start_row + LOT_SIZE - 1, ws.max_row)
            print(f"➡️ Traitement du lot {lot_index + 1}/{total_lots} (lignes {start_row}-{end_row})")

            for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                try:
                    nom, prenom, telephone, email, formation, mode = row[:6]
                    if not email or not telephone:
                        continue

                    email = email.strip().lower()
                    telephone = str(telephone).strip().replace(" ", "")
                    if telephone.startswith("0"):
                        telephone = "+33" + telephone[1:]

                    cur.execute("SELECT id FROM parcoursup_candidats WHERE email=? OR telephone=?", (email, telephone))
                    if cur.fetchone():
                        duplicates += 1
                        continue

                    cid = str(uuid.uuid4())
                    now = datetime.now().isoformat()

                    cur.execute("""
                        INSERT INTO parcoursup_candidats 
                        (id, nom, prenom, telephone, email, formation, mode, mail_ok, sms_ok, statut, logs, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, 'En attente de candidature', '[]', ?)
                    """, (cid, nom, prenom, telephone, email, formation, mode, now))
                    imported += 1

                                     # === Envoi du mail ===
                    mail_body = mail_html(
                        "parcoursup_import",
                        prenom=prenom,
                        bts_label=formation,
                        lien_espace="https://inscriptionsbts.onrender.com/"
                    )
                    if send_mail(email, "Votre candidature Parcoursup – Intégrale Academy", mail_body):
                        mails_sent += 1
                        cur.execute("UPDATE parcoursup_candidats SET mail_ok=1 WHERE id=?", (cid,))
                        cur.execute("UPDATE parcoursup_candidats SET logs=json_insert(logs, '$[#]', json_object('type', 'mail', 'dest', ?, 'date', ?)) WHERE id=?", (email, now, cid))

                    # === Envoi du SMS ===
                    sms_body = sms_text(
                        "parcoursup_import",
                        prenom=prenom,
                        bts_label=formation,
                        lien_espace="https://inscriptionsbts.onrender.com/"
                    )
                    sms_id = send_sms_brevo(telephone, sms_body)
                    if sms_id:
                        sms_sent += 1
                        cur.execute("UPDATE parcoursup_candidats SET sms_ok=1 WHERE id=?", (cid,))
                        cur.execute("UPDATE parcoursup_candidats SET logs=json_insert(logs, '$[#]', json_object('type', 'sms', 'dest', ?, 'id', ?, 'date', ?)) WHERE id=?", (telephone, str(sms_id), now, cid))



                    row_index += 1

                except Exception as e:
                    errors += 1
                    print(f"⚠️ Erreur ligne {row_index}: {e}")

            conn.commit()
            print(f"✅ Lot {lot_index + 1}/{total_lots} terminé — total partiel : {imported} importés, {mails_sent} mails, {sms_sent} SMS")
            time.sleep(2)  # 💤 pause 2 secondes entre lots

    finally:
        conn.close()
        os.remove(temp_path)

    flash(f"{imported} importés — {mails_sent} mails — {sms_sent} SMS — {duplicates} doublons — {errors} erreurs.", "success")
    print("🏁 Import terminé avec succès.")
    return redirect(url_for("parcoursup.dashboard"))

# =====================================================
# 🕵️‍♂️ VÉRIFICATION DU FICHIER EXCEL (AVANT IMPORT)
# =====================================================
@bp_parcoursup.route("/parcoursup/check", methods=["POST"])
def check_file():
    if "file" not in request.files:
        flash("Aucun fichier sélectionné", "error")
        return redirect(url_for("parcoursup.dashboard"))

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        flash("Format de fichier non supporté. Utilisez un fichier .xlsx", "error")
        return redirect(url_for("parcoursup.dashboard"))

    temp_path = os.path.join(DATA_DIR, secure_filename(file.filename))
    file.save(temp_path)
    wb = load_workbook(temp_path)
    ws = wb.active

    erreurs = []
    corrections = 0
    lignes_supprimees = 0
    rows_to_delete = []
    ligne_num = 2

    for row_cells in ws.iter_rows(min_row=2):
        telephone = row_cells[2].value if len(row_cells) > 2 else None
        email = row_cells[3].value if len(row_cells) > 3 else None
        mode = row_cells[5].value if len(row_cells) > 5 else None

        tel = str(telephone or "").strip().replace(" ", "")
        tel_clean = _clean_phone(telephone)
        if tel_clean != tel and len(row_cells) > 2:
            row_cells[2].value = tel_clean
            corrections += 1

        mail = (email or "").strip().lower()
        if len(row_cells) > 3 and mail != str(email or ""):
            row_cells[3].value = mail
            corrections += 1
        
        tel_ok = bool(re.match(r"^(?:\+33|0)[1-9]\d{8}$", tel_clean))
        mail_ok = ("@" in mail and "." in mail)
        ligne_vide = (not tel_clean and not mail)

        if not ligne_vide and (not tel_ok or not mail_ok):
            if not tel_ok:
                erreurs.append(f"Ligne {ligne_num} : téléphone invalide ({tel})")
            if not mail_ok:
                erreurs.append(f"Ligne {ligne_num} : e-mail invalide ({mail})")
            rows_to_delete.append(ligne_num)

        mode_clean = _clean_mode(mode)
        if mode_clean != mode and len(row_cells) > 5:
            row_cells[5].value = mode_clean
            corrections += 1

        ligne_num += 1

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)
        lignes_supprimees += 1

    os.remove(temp_path)
    if erreurs:
        msg = f"❌ {len(erreurs)} erreur(s) détectée(s) :<br>" + "<br>".join(erreurs[:20])
        if len(erreurs) > 20:
            msg += f"<br>… et {len(erreurs) - 20} autres lignes à corriger."
        cleaned_token = str(uuid.uuid4())
        cleaned_name = f"parcoursup_nettoye_{cleaned_token}.xlsx"
        cleaned_path = os.path.join(CLEANED_DIR, cleaned_name)
        wb.save(cleaned_path)
        clean_url = url_for("parcoursup.download_cleaned_file", token=cleaned_token)
        msg += (
            f"<br><br>🧹 Un nettoyage automatique a été préparé ({corrections} correction(s), {lignes_supprimees} ligne(s) supprimée(s)). "
            f"<a href='{clean_url}' target='_blank' rel='noopener'>Télécharger le fichier nettoyé</a>."
        )
        flash(Markup(msg), "error")
    else:
        flash("✅ Aucun problème détecté : le fichier est prêt à être importé.", "success")

    return redirect(url_for("parcoursup.dashboard"))

@bp_parcoursup.route("/parcoursup/cleaned/<token>", methods=["GET"])
def download_cleaned_file(token):
    safe_token = secure_filename(token)
    pattern = os.path.join(CLEANED_DIR, f"parcoursup_nettoye_{safe_token}.xlsx")
    if not os.path.exists(pattern):
        flash("Fichier nettoyé introuvable (lien expiré).", "error")
        return redirect(url_for("parcoursup.dashboard"))
    return send_file(
        pattern,
        as_attachment=True,
        download_name="parcoursup_nettoye.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# 📩 RELANCER AUTOMATIQUEMENT LES NON OUVERTS APRÈS 48H
# =====================================================
@bp_parcoursup.route("/parcoursup/relancer-non-ouverts", methods=["POST"])
def relancer_non_ouverts():
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, prenom, email, telephone, formation, logs FROM parcoursup_candidats")
    rows = cur.fetchall()
    relances = 0
    now = datetime.now().isoformat()

    for r in rows:
        try:
            logs = json.loads(r["logs"] or "[]")
            mail_log = next((l for l in logs if l.get("type") == "mail"), None)
            opened_or_clicked = any(
                l.get("event") in ["opened", "click", "unique_opened"]
                for l in logs if l.get("type") == "mail_status"
            )

            if mail_log and not opened_or_clicked:
                mail_date = datetime.fromisoformat(mail_log["date"])
                diff_hours = (datetime.now() - mail_date).total_seconds() / 3600

                if diff_hours > 48:
                    prenom = r["prenom"] or ""
                    formation = r["formation"] or ""
                    email = r["email"] or ""
                    tel = r["telephone"] or ""

                    # --- MAIL ---
                    html = f"""
                    <p>Bonjour {prenom},</p>
                    <p>Nous n’avons pas encore reçu votre confirmation Parcoursup pour le BTS <b>{formation}</b>.</p>
                    <p>Merci de compléter votre pré-inscription ici :</p>
                    <p><a href='https://inscriptionsbts.onrender.com/'><b>👉 Formulaire de pré-inscription</b></a></p>
                    <p>À bientôt,<br><b>L’équipe Intégrale Academy</b></p>
                    """
                    send_mail(email, "Relance – Intégrale Academy", html)

                    # --- SMS ---
                    msg = f"Bonjour {prenom}, merci de finaliser votre pré-inscription BTS {formation} ici 👉 inscriptionsbts.onrender.com"
                    send_sms_brevo(tel, msg)

                    # --- Log ---
                    cur.execute(
                        "UPDATE parcoursup_candidats SET logs=json_insert(logs, '$[#]', json_object('type','relance_auto','date',?)) WHERE id=?",
                        (now, r["id"])
                    )
                    relances += 1

        except Exception as e:
            print(f"⚠️ Erreur relance auto: {e}")
            continue

    conn.commit()
    conn.close()

    flash(f"{relances} relances automatiques envoyées.", "success")
    return redirect(url_for("parcoursup.dashboard"))


# =====================================================
# 🗑️ SUPPRIMER UN CANDIDAT
# =====================================================
@bp_parcoursup.route("/parcoursup/delete/<cid>", methods=["POST"])
def delete_candidat(cid):
    """Supprime un candidat Parcoursup depuis le tableau d'administration."""
    conn = db()
    cur = conn.cursor()
    cur.execute("DELETE FROM parcoursup_candidats WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    flash("Candidature supprimée avec succès.", "success")
    return redirect(url_for("parcoursup.dashboard"))


# =====================================================
# 📬 VÉRIFIER LES STATUTS SMS
# =====================================================
@bp_parcoursup.route("/parcoursup/check-sms", methods=["POST"])
def check_sms_status_all():
    BREVO_KEY = os.getenv("BREVO_API_KEY")
    if not BREVO_KEY:
        flash("BREVO_API_KEY manquant dans les variables d'environnement.", "error")
        return redirect(url_for("parcoursup.dashboard"))

    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, logs FROM parcoursup_candidats WHERE sms_ok=1")
    rows = cur.fetchall()

    headers = {"api-key": BREVO_KEY}
    delivered = failed = pending = 0

    # ✅ Fonction interne pour vérifier le statut SMS
    def last_event(message_id: str):
        """Retourne le dernier statut du SMS via Brevo (version compatible 2025)."""
        headers = {"api-key": BREVO_KEY}
        try:
            # ✅ Nouvelle route correcte
            url = f"https://api.brevo.com/v3/transactionalSMS/messages/{message_id}"
            r = requests.get(url, headers=headers, timeout=15)

            print("📦 Réponse complète Brevo:", r.text)


            if not r.ok:
                current_app.logger.warning(f"⚠️ Requête non OK ({r.status_code}): {r.text[:300]}")
                return "unknown"

            data = r.json()
            events = data.get("events") or data.get("messages") or []
            if not events:
                current_app.logger.info("ℹ️ Aucun évènement trouvé pour ce message_id.")
                return "unknown"

            # 🧩 On prend le dernier évènement (souvent 'delivered', 'failed', 'queued'…)
            last = events[-1]
            evt = last.get("event") or last.get("status") or "unknown"
            current_app.logger.info(f"📩 Statut SMS {message_id}: {evt}")
            return evt

        except Exception as e:
            current_app.logger.exception(f"❌ Exception last_event(): {e}")
            return "unknown"

    # 🔁 Boucle principale
    for r in rows:
        try:
            raw_logs = r["logs"] or "[]"
            logs = json.loads(raw_logs) if raw_logs else []
            sms_log = next((l for l in logs if l.get("type") == "sms" and l.get("id")), None)
            if not sms_log:
                continue

            evt = last_event(sms_log["id"])
            now = datetime.now().isoformat()

            if evt == "delivered":
                delivered += 1
                cur.execute("""
                    UPDATE parcoursup_candidats
                    SET sms_ok=1, logs=json_insert(logs, '$[#]',
                        json_object('type','sms_status','event',?,'date',?))
                    WHERE id=?""", (evt, now, r["id"]))
            elif evt == "failed":
                failed += 1
                cur.execute("""
                    UPDATE parcoursup_candidats
                    SET sms_ok=0, logs=json_insert(logs, '$[#]',
                        json_object('type','sms_status','event',?,'date',?))
                    WHERE id=?""", (evt, now, r["id"]))
            else:
                pending += 1
                cur.execute("""
                    UPDATE parcoursup_candidats
                    SET logs=json_insert(logs, '$[#]',
                        json_object('type','sms_status','event',?,'date',?))
                    WHERE id=?""", (evt or "unknown", now, r["id"]))

        except Exception as e:
            current_app.logger.exception(f"❌ boucle check_sms_status: {e}")
        finally:
            time.sleep(0.15)

    conn.commit()
    conn.close()
    flash(f"SMS livrés ✅ {delivered} — échoués ❌ {failed} — en attente ⏳ {pending}", "success")
    return redirect(url_for("parcoursup.dashboard"))






@bp_parcoursup.route("/parcoursup/logs/<cid>")
def get_logs(cid):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT logs FROM parcoursup_candidats WHERE id=?", (cid,))
    row = cur.fetchone()
    conn.close()

    if not row:
        return jsonify([])

    try:
        logs = json.loads(row["logs"])
    except Exception:
        logs = []
    return jsonify(logs)

# =====================================================
# 🔁 RELANCER MANUELLEMENT UN CANDIDAT (MAIL + SMS)
# =====================================================
@bp_parcoursup.route("/parcoursup/relancer/<cid>", methods=["POST"])
def relancer_individuel(cid):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT prenom, email, telephone, formation FROM parcoursup_candidats WHERE id=?", (cid,))
    r = cur.fetchone()

    if not r:
        flash("Candidat introuvable.", "error")
        conn.close()
        return redirect(url_for("parcoursup.dashboard"))

    prenom = r["prenom"] or ""
    formation = r["formation"] or ""
    email = (r["email"] or "").strip()
    tel = (r["telephone"] or "").strip()
    now = datetime.now().isoformat()

    try:
        # 🔗 Lien espace candidat (affiché dans le mail et SMS)
        base_url = os.getenv("BASE_URL", "https://inscriptionsbts.onrender.com").rstrip("/")
        lien_espace = f"{base_url}/espace/{cid}"

        # 📬 Génération des contenus depuis les templates
        mail_body = mail_html("parcoursup_relance", prenom=prenom, bts_label=formation, lien_espace=lien_espace)
        sms_body = sms_text("parcoursup_relance", prenom=prenom, bts_label=formation, lien_espace=lien_espace)


        # 📧 Envoi du mail
        mail_ok = False
        try:
            send_mail(email, "📩 Relance — Votre candidature Intégrale Academy", mail_body)
            mail_ok = True
        except Exception as e:
            print(f"⚠️ Erreur mail: {e}")

        # 📱 Envoi du SMS
        sms_ok = False
        try:
            send_sms_brevo(tel, sms_body)
            sms_ok = True
        except Exception as e:
            print(f"⚠️ Erreur SMS: {e}")

        # 🕓 Ajout dans les logs
        cur.execute("""
            UPDATE parcoursup_candidats
            SET logs = json_insert(
                logs, '$[#]',
                json_object(
                    'type', 'relance_manuelle',
                    'modele', 'candidature_validee',
                    'date', ?
                )
            )
            WHERE id = ?
        """, (now, cid))
        conn.commit()

        flash(f"Relance envoyée à {prenom} ✅ (mail:{'OK' if mail_ok else '❌'}, sms:{'OK' if sms_ok else '❌'})", "success")

    except Exception as e:
        flash(f"Erreur lors de la relance : {e}", "error")

    conn.close()
    return redirect(url_for("parcoursup.dashboard"))


    prenom = r["prenom"] or ""
    formation = r["formation"] or ""
    email = (r["email"] or "").strip()
    tel = (r["telephone"] or "").strip()
    now = datetime.now().isoformat()

    try:
        # --- MAIL ---
        html = f"""
        <p>Bonjour {prenom},</p>
        <p>Nous vous relançons concernant votre candidature Parcoursup pour le BTS <b>{formation}</b>.</p>
        <p>Merci de compléter votre pré-inscription ici :</p>
        <p><a href='https://inscriptionsbts.onrender.com/'><b>👉 Formulaire de pré-inscription</b></a></p>
        <p>À bientôt,<br><b>L’équipe Intégrale Academy</b></p>
        """
        send_mail(email, "Relance manuelle – Intégrale Academy", html)

        # --- SMS ---
        msg = f"Bonjour {prenom}, merci de finaliser votre pré-inscription BTS {formation} ici 👉 inscriptionsbts.onrender.com"
        send_sms_brevo(tel, msg)

        # --- Log en BDD ---
        cur.execute(
            "UPDATE parcoursup_candidats SET logs=json_insert(logs, '$[#]', json_object('type','relance_manuelle','date',?)) WHERE id=?",
            (now, cid)
        )
        conn.commit()
        flash(f"Relance envoyée à {prenom}.", "success")

    except Exception as e:
        flash(f"Erreur lors de la relance : {e}", "error")

    conn.close()
    return redirect(url_for("parcoursup.dashboard"))


# =====================================================
# 📬 WEBHOOK SMS BREVO — MISE À JOUR STATUTS SMS
# =====================================================
@bp_parcoursup.route("/brevo-sms-webhook", methods=["POST"])
def brevo_sms_webhook():
    """Réception des notifications SMS de Brevo (sent, delivered, failed)."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        print(f"📩 Webhook Brevo reçu : {data}")

        # On récupère les champs selon les variantes possibles
        msg_status = data.get("msg_status") or data.get("event") or ""
        message_id = str(data.get("messageId") or data.get("message_id") or "").strip()
        phone = str(data.get("to") or data.get("recipient") or "").replace(" ", "").replace("0033", "+33")
        now = datetime.now().isoformat()

        if not message_id or not msg_status:
            print("⚠️ Webhook ignoré : données incomplètes.")
            return jsonify({"status": "ignored"}), 400

        # Connexion BDD
        conn = db()
        cur = conn.cursor()

        # On recherche le candidat correspondant à ce messageId
        cur.execute("SELECT id, logs FROM parcoursup_candidats")
        rows = cur.fetchall()
        found = False
        for r in rows:
            logs = json.loads(r["logs"] or "[]")
            for l in logs:
                if str(l.get("id")) == message_id:
                    found = True
                    cur.execute("""
                        UPDATE parcoursup_candidats
                        SET sms_ok = CASE
                            WHEN ? = 'delivered' THEN 1
                            WHEN ? IN ('failed', 'rejected', 'error') THEN 0
                            ELSE sms_ok END,
                        logs = json_insert(
                            logs, '$[#]',
                            json_object('type', 'sms_status', 'event', ?, 'date', ?)
                        )
                        WHERE id = ?
                    """, (msg_status, msg_status, msg_status, now, r["id"]))
                    conn.commit()
                    print(f"✅ Statut SMS mis à jour : {msg_status} pour {phone}")
                    break
            if found:
                break

        conn.close()

        if not found:
            print(f"⚠️ Aucun candidat trouvé pour message_id={message_id}")
            return jsonify({"status": "not_found"}), 404

        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print(f"❌ Erreur traitement webhook SMS : {e}")
        return jsonify({"status": "error", "error": str(e)}), 500

@bp_parcoursup.route("/brevo-mail-webhook", methods=["POST"])
def brevo_mail_webhook():
    """Webhook Brevo pour suivi des e-mails Parcoursup."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        print(f"📩 Webhook Brevo MAIL reçu : {data}")

        message_id = str(data.get("messageId") or data.get("message-id") or "").strip()
        event = data.get("event") or data.get("type") or ""
        email = (data.get("email") or "").lower().strip()
        now = datetime.now().isoformat()

        if not message_id or not event:
            print("⚠️ Webhook mail ignoré : données incomplètes.")
            return jsonify({"status": "ignored"}), 400

        conn = db()
        cur = conn.cursor()
        cur.execute("SELECT id, logs FROM parcoursup_candidats")
        rows = cur.fetchall()
        found = False

        for r in rows:
            logs = json.loads(r["logs"] or "[]")
            for l in logs:
                if str(l.get("id")) == message_id or l.get("dest") == email:
                    found = True
                    cur.execute("""
                        UPDATE parcoursup_candidats
                        SET mail_ok = CASE
                            WHEN ? IN ('delivered', 'opened', 'clicked') THEN 1
                            ELSE mail_ok END,
                        logs = json_insert(
                            logs, '$[#]',
                            json_object('type','mail_status','event',?, 'date',?)
                        )
                        WHERE id = ?
                    """, (event, event, now, r["id"]))
                    conn.commit()
                    print(f"✅ Statut mail mis à jour : {event} pour {email}")
                    break
            if found:
                break

        conn.close()
        if not found:
            print(f"⚠️ Aucun candidat trouvé pour mail_id={message_id}")
            return jsonify({"status": "not_found"}), 404

        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print(f"❌ Erreur traitement webhook MAIL : {e}")
        return jsonify({"status": "error", "error": str(e)}), 500

































